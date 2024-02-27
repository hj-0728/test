from io import BytesIO
from typing import List

from infra_basic.transaction import Transaction
from infra_object_storage.model.file_info_model import FileExtInfoModel
from infra_object_storage.service.object_storage_service import ObjectStorageService
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from edu_binshi.data.query_params.student_query_params import StudentPageQueryParams
from edu_binshi.model.view.student_user_page_vm import StudentUserPageVm
from edu_binshi.service.student_service import StudentService


class AppStudentService:
    def __init__(
        self,
        student_service: StudentService,
        object_storage_service: ObjectStorageService,
    ):
        self.__student_service = student_service
        self.__object_storage_service = object_storage_service

    def export_student_user(
        self, params: StudentPageQueryParams, transaction: Transaction
    ) -> FileExtInfoModel:
        """
        导出学生用户
        :param params:
        :param transaction:
        :return:
        """

        params.page_size = 1000000
        params.page_index = 0

        student_user_page = self.__student_service.get_student_user_page(params=params)

        student_init_password = self.__student_service.get_student_init_password()

        file_io = self.prepare_export_student_user_excel(
            student_user_data=student_user_page.data, student_init_password=student_init_password
        )

        file_info = self.__object_storage_service.upload_file(
            file_name=f"学生账号.xlsx",
            file_blob=file_io,
            transaction=transaction,
        )
        return file_info

    @staticmethod
    def prepare_export_student_user_excel(
        student_user_data: List[StudentUserPageVm], student_init_password: str
    ):
        """

        :param student_user_data:
        :return:
        """
        work_book = Workbook()

        sheet = work_book.active

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        title_font = Font(name="微软雅黑", size=12)
        align = Alignment(horizontal="center", vertical="center", wrapText=True)

        title_list = ["姓名", "班级", "账号", "是否初始密码", "状态"]

        sheet.merge_cells("A1:E1")
        cell = sheet["A1"]
        cell.value = f"""
        注意：
        1.平台初始密码为：{student_init_password}，请提示学生及时修改。
        2.若管理员/老师为学生重置过密码或学生已修改密码并登录平台，在最后一列“是否初始密码”会显示“否”，请学生用重置后或自行修改后的密码登录。
        """
        red_font = Font(color=Color("FF0000"))
        cell.font = red_font
        # 设置行高
        sheet.row_dimensions[1].height = 90
        # 设置文字自动换行
        cell.alignment = Alignment(wrap_text=True)

        row = 2
        for i in range(1, len(title_list) + 1):
            sheet.cell(row, i).value = title_list[i - 1]
            sheet.cell(row, i).border = thin_border
            sheet.cell(row, i).font = title_font
            sheet.cell(row, i).fill = PatternFill("solid", fgColor="C5D9F1")
            sheet.cell(row, i).alignment = align
            col_letter = get_column_letter(i)
            width = 26
            if i == 2:
                width = 40
            sheet.column_dimensions[col_letter].width = width
        row += 1
        font = Font(name="等线", size=11)
        for student_user in student_user_data:
            if student_user.user:
                sheet.cell(row, 1).value = student_user.student_name
                sheet.cell(row, 2).value = f"{student_user.school_class}"
                sheet.cell(row, 3).value = student_user.user.name
                sheet.cell(row, 4).value = "是" if student_user.user.password_reset else "否"
                sheet.cell(row, 5).value = "已启用" if student_user.user.is_activated else "未启用"
                for i in range(1, 6):
                    sheet.cell(row, i).font = font
                    sheet.cell(row, i).border = thin_border
                if not student_user.user.is_activated:
                    sheet.cell(row, 5).font = red_font
                row += 1
        excel_stream = BytesIO()
        work_book.save(excel_stream)
        return excel_stream.getvalue()
